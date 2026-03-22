#!/usr/bin/env python3
"""
JudgeIQ Judging Analytics Engine
Analyzes competition judge scoring data and produces per-division Excel sheets + JSON output.

Usage: python analyze.py input.json output_dir/
"""

import json
import sys
import os
import math
from pathlib import Path
from collections import defaultdict

import numpy as np
import pandas as pd
from scipy import stats
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Format Detection & Parsing
# ---------------------------------------------------------------------------

def is_tabroom_tournament(obj) -> bool:
    return isinstance(obj, dict) and ("categories" in obj or "judges" in obj)


def parse_tabroom_tournament(data: dict) -> tuple:
    """
    Parse a single Tabroom tournament export.
    Returns (entries, meta, judge_id_map).
    judge_id_map: {judge_full_name: judge_id}
    """
    def _build_judge_map(judge_list):
        """Build {str(id): 'First Last'} and {'First Last': str(id)} from a judges array."""
        jmap, jid_map = {}, {}
        for j in (judge_list or []):
            if not isinstance(j, dict):
                continue
            jid = str(j.get("id", "")).strip()
            name = f"{j.get('first', '')} {j.get('last', '')}".strip()
            if jid and name:
                jmap[jid] = name
                jid_map[name] = jid
        return jmap, jid_map

    # Build judge map — check top-level first, then each category
    judge_map, judge_id_map = _build_judge_map(data.get("judges", []))

    ballot_records = []
    divisions_seen = []

    for category in data.get("categories", []):
        div_name = (category.get("name") or category.get("abbr") or "Unknown").strip()

        # Merge in any judges declared at the category level
        cat_jmap, cat_jid_map = _build_judge_map(category.get("judges", []))
        cat_jmap.update(judge_map)      # top-level takes precedence if duplicate
        cat_jid_map.update(judge_id_map)

        # Collect rounds from all possible paths
        rounds = []
        # Path 1: category → events → rounds
        for event in category.get("events", []):
            rounds.extend(event.get("rounds", []))
        # Path 2: category → rounds (direct, no events wrapper)
        if not rounds:
            rounds.extend(category.get("rounds", []))

        div_had_ballots = False
        for round_ in rounds:
            for section in (round_.get("sections") or round_.get("panels") or []):
                for ballot in section.get("ballots", []):
                    # Skip explicit byes/forfeits (int 1, bool True, or truthy)
                    if ballot.get("bye") or ballot.get("forfeit"):
                        continue

                    # Normalize judge ID to string for lookup
                    raw_judge = ballot.get("judge")
                    if isinstance(raw_judge, dict):
                        judge_id = str(raw_judge.get("id", "")).strip()
                    else:
                        judge_id = str(raw_judge).strip() if raw_judge is not None else ""
                    judge_name = cat_jmap.get(judge_id)
                    if not judge_name:
                        continue

                    # Competitor: entry is often a bare int/str ID
                    comp_raw = ballot.get("competitor") or ballot.get("entry")
                    if isinstance(comp_raw, dict):
                        comp_id = str(comp_raw.get("id") or "")
                        comp_name = str(comp_raw.get("name") or comp_raw.get("code") or comp_id)
                    elif comp_raw is not None:
                        comp_id = str(comp_raw)
                        comp_name = comp_id
                    else:
                        comp_id = str(ballot.get("entry_id") or "")
                        comp_name = comp_id

                    # Speaker points: collect each individual score with tag "point"
                    point_scores = []
                    for s in (ballot.get("scores") or []):
                        if not isinstance(s, dict):
                            continue
                        if s.get("tag") == "point":
                            try:
                                point_scores.append(float(s["value"]))
                            except (TypeError, ValueError, KeyError):
                                pass

                    # Skip ballots with no points or all-zero (bye proxy)
                    if not point_scores or all(v == 0 for v in point_scores):
                        continue

                    div_had_ballots = True
                    for pt in point_scores:
                        ballot_records.append({
                            "competitor_id": comp_id,
                            "competitor_name": comp_name,
                            "division": div_name,
                            "judge_name": judge_name,
                            "score": pt,
                        })
        if div_had_ballots and div_name not in divisions_seen:
            divisions_seen.append(div_name)

    entries = _aggregate_ballot_records(ballot_records)
    meta = {
        "tournament_name": data.get("name", "Unknown Tournament"),
        "judge_count": len(judge_map),
        "divisions": divisions_seen,
        "total_ballots": len(ballot_records),
    }
    return entries, meta, judge_id_map


def _aggregate_ballot_records(ballot_records: list) -> list:
    comp_info = {}
    scores_by_comp_judge = defaultdict(list)
    for record in ballot_records:
        cid = record["competitor_id"]
        if cid not in comp_info:
            comp_info[cid] = {"name": record["competitor_name"], "division": record["division"]}
        scores_by_comp_judge[(cid, record["judge_name"])].append(record["score"])

    result = []
    for comp_id, info in comp_info.items():
        judge_scores = {
            jname: sc
            for (cid, jname), sc in scores_by_comp_judge.items()
            if cid == comp_id
        }
        result.append({
            "competitor_id": comp_id,
            "competitor_name": info["name"],
            "division": info["division"],
            "judge_scores": judge_scores,
            "final_placement": None,
        })
    return result


def parse_unknown_format(data) -> tuple:
    def find_score_maps(obj, depth=0):
        if depth > 12:
            return []
        found = []
        if isinstance(obj, dict):
            numeric = {k: v for k, v in obj.items()
                       if isinstance(v, (int, float)) and not isinstance(v, bool) and 0 <= v <= 300}
            non_numeric = {k: v for k, v in obj.items() if k not in numeric}
            if len(numeric) >= 2:
                comp_id = str(non_numeric.get("id") or non_numeric.get("competitor_id") or "")
                comp_name = str(non_numeric.get("name") or non_numeric.get("competitor_name") or comp_id)
                division = str(non_numeric.get("division") or non_numeric.get("category") or "Unknown")
                placement = non_numeric.get("final_placement") or non_numeric.get("placement")
                found.append({
                    "competitor_id": comp_id,
                    "competitor_name": comp_name,
                    "division": division,
                    "judge_scores": {str(k): float(v) for k, v in numeric.items()},
                    "final_placement": int(placement) if placement is not None else None,
                })
            for v in obj.values():
                found.extend(find_score_maps(v, depth + 1))
        elif isinstance(obj, list):
            for item in obj:
                found.extend(find_score_maps(item, depth + 1))
        return found

    entries = find_score_maps(data)
    seen = set()
    unique = []
    for e in entries:
        key = (e["competitor_id"], frozenset(e["judge_scores"].items()))
        if key not in seen:
            seen.add(key)
            unique.append(e)

    all_judges = set()
    for e in unique:
        all_judges.update(e["judge_scores"].keys())

    return unique, {"judge_count": len(all_judges), "entry_count": len(unique)}, {}


def load_data(filepath: str) -> tuple:
    """
    Load and auto-detect JSON format.
    Returns (entries, format_info, judge_id_map).
    format_info["judge_id_map"] = {judge_full_name: judge_id}
    """
    for enc in ("utf-8", "latin-1", "utf-8-sig"):
        try:
            with open(filepath, "r", encoding=enc) as f:
                data = json.load(f)
            break
        except (UnicodeDecodeError, json.JSONDecodeError):
            data = None
    if data is None:
        raise ValueError("Could not decode the JSON file. Ensure it is valid UTF-8 or Latin-1 encoded.")

    # ── Tabroom single tournament ──────────────────────────────────────────
    if is_tabroom_tournament(data):
        entries, meta, judge_id_map = parse_tabroom_tournament(data)
        if not entries:
            raise ValueError(
                f"Tabroom tournament '{meta.get('tournament_name', '')}' detected but no "
                "scorable ballots found. Ensure rounds have speaker point scores (tag='point') "
                "that are not byes or forfeits."
            )
        format_info = {
            "format": "tabroom",
            "label": f"Tabroom tournament detected: {meta['tournament_name']}",
            "tournament_name": meta["tournament_name"],
            "judge_count": meta["judge_count"],
            "division_count": len(meta["divisions"]),
            "divisions": meta["divisions"],
            "total_ballots": meta["total_ballots"],
            "judge_id_map": judge_id_map,
        }
        return entries, format_info

    # ── Tabroom multi-tournament array ─────────────────────────────────────
    if isinstance(data, list) and data and all(is_tabroom_tournament(d) for d in data):
        all_entries, all_divisions, all_id_map = [], [], {}
        total_ballots, tournament_names = 0, []

        for tournament in data:
            t_entries, t_meta, t_id_map = parse_tabroom_tournament(tournament)
            all_entries.extend(t_entries)
            total_ballots += t_meta["total_ballots"]
            tournament_names.append(t_meta["tournament_name"])
            all_id_map.update(t_id_map)
            for div in t_meta["divisions"]:
                if div not in all_divisions:
                    all_divisions.append(div)

        if not all_entries:
            raise ValueError("Multi-tournament Tabroom export detected but no scorable ballots found.")

        names_str = ", ".join(tournament_names[:3])
        if len(tournament_names) > 3:
            names_str += f" (+{len(tournament_names) - 3} more)"

        all_judges = set()
        for e in all_entries:
            all_judges.update(e["judge_scores"].keys())

        format_info = {
            "format": "tabroom_multi",
            "label": f"Multi-tournament Tabroom export: {names_str}",
            "tournament_names": tournament_names,
            "judge_count": len(all_judges),
            "division_count": len(all_divisions),
            "divisions": all_divisions,
            "total_ballots": total_ballots,
            "judge_id_map": all_id_map,
        }
        return all_entries, format_info

    # ── Legacy array with judge_scores ────────────────────────────────────
    if isinstance(data, list) and data and all(isinstance(d, dict) and "judge_scores" in d for d in data):
        all_judges = set()
        for e in data:
            all_judges.update(e.get("judge_scores", {}).keys())
        format_info = {
            "format": "legacy_array",
            "label": f"Custom format detected — found {len(all_judges)} judges with scoring data",
            "judge_count": len(all_judges),
            "entry_count": len(data),
            "judge_id_map": {},
        }
        return data, format_info

    # ── Unknown: heuristic scan ────────────────────────────────────────────
    entries, meta, judge_id_map = parse_unknown_format(data)
    if not entries:
        raise ValueError(
            "Could not extract scoring data from this JSON. "
            "Supported formats: Tabroom tournament export (object with 'categories'/'judges'), "
            "Tabroom multi-tournament array, or an array of objects with a 'judge_scores' key."
        )
    format_info = {
        "format": "unknown",
        "label": f"Custom format detected — found {meta['judge_count']} judges with scoring data",
        "judge_count": meta["judge_count"],
        "entry_count": meta["entry_count"],
        "judge_id_map": {},
    }
    return entries, format_info


# ---------------------------------------------------------------------------
# Score Matrix
# ---------------------------------------------------------------------------

def extract_judge_names(data: list) -> list:
    names = set()
    for entry in data:
        for j, scores in entry.get("judge_scores", {}).items():
            if isinstance(scores, list) and scores:
                names.add(j)
            elif not isinstance(scores, list) and scores is not None:
                names.add(j)
    return sorted(names)


def build_score_matrix(data: list, judges: list) -> pd.DataFrame:
    rows = []
    for entry in data:
        row = {
            "competitor_id": entry.get("competitor_id", ""),
            "competitor_name": entry.get("competitor_name", ""),
            "division": entry.get("division", ""),
            "final_placement": entry.get("final_placement", None),
        }
        scores = entry.get("judge_scores", {})
        for judge in judges:
            row[judge] = scores.get(judge, np.nan)
        rows.append(row)
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Per-Division Judge Statistics (new format)
# ---------------------------------------------------------------------------

STAT_COLUMNS = [
    "Severity_Rank", "Judge_ID", "Judge_Full_Name", "Number_of_Scores",
    "Mean", "Z_Severity_Index", "Absolute_Severity", "Bias_Direction",
    "Mann_Whitney_U", "p_value",
]


def compute_division_judge_stats(
    entries: list, judges: list, judge_id_map: dict
) -> pd.DataFrame:
    """
    Compute per-judge stats for a single division's entries.
    Returns a DataFrame with STAT_COLUMNS, sorted by Severity_Rank ascending.
    """
    # Build per-judge score arrays by flattening individual speaker point scores
    judge_scores_dict = {}
    for j in judges:
        scores = []
        for e in entries:
            js = e.get("judge_scores", {}).get(j)
            if js is not None:
                if isinstance(js, list):
                    scores.extend(js)
                else:
                    scores.append(js)
        if scores:
            judge_scores_dict[j] = np.array(scores)

    valid_judges = list(judge_scores_dict.keys())
    if len(valid_judges) < 2:
        return pd.DataFrame(columns=STAT_COLUMNS)

    judge_means = {j: float(v.mean()) for j, v in judge_scores_dict.items()}

    means_arr = np.array(list(judge_means.values()))
    mean_of_means = float(means_arr.mean())
    std_of_means = float(means_arr.std(ddof=1)) if len(means_arr) > 1 else 0.0

    records = []
    for judge in valid_judges:
        scores = judge_scores_dict[judge]
        n = len(scores)
        mean_score = judge_means[judge]

        # Z_Severity_Index: how many std-devs of judge-mean spread this judge is from center
        z_severity = (mean_score - mean_of_means) / std_of_means if std_of_means > 0 else 0.0
        abs_severity = abs(z_severity)
        bias_direction = "High-Side" if z_severity >= 0 else "Low-Side"

        # Mann-Whitney U: this judge's scores vs all other judges' scores pooled
        other_scores = np.concatenate([
            v for jn, v in judge_scores_dict.items() if jn != judge
        ])

        if n >= 3 and len(other_scores) >= 3:
            try:
                u_stat, p_val = stats.mannwhitneyu(scores, other_scores, alternative="two-sided")
            except ValueError:
                u_stat, p_val = np.nan, np.nan
        else:
            u_stat, p_val = np.nan, np.nan

        jid = judge_id_map.get(judge, "")

        records.append({
            "Judge_ID": jid,
            "Judge_Full_Name": judge,
            "Number_of_Scores": int(n),
            "Mean": round(mean_score, 2),
            "Z_Severity_Index": round(z_severity, 4),
            "Absolute_Severity": round(abs_severity, 4),
            "Bias_Direction": bias_direction,
            "Mann_Whitney_U": round(float(u_stat), 1) if not (isinstance(u_stat, float) and math.isnan(u_stat)) else None,
            "p_value": float(p_val) if not (isinstance(p_val, float) and math.isnan(p_val)) else None,
        })

    df_result = pd.DataFrame(records)
    df_result = df_result.sort_values("Absolute_Severity", ascending=False).reset_index(drop=True)
    df_result.insert(0, "Severity_Rank", range(1, len(df_result) + 1))
    return df_result[STAT_COLUMNS]


# ---------------------------------------------------------------------------
# Division Summary Stats (for dashboard)
# ---------------------------------------------------------------------------

def compute_division_summary(entries: list, judges: list) -> dict:
    """Compute overall division-level stats for the dashboard."""
    all_scores = []
    active_judges = []
    for j in judges:
        scores = []
        for e in entries:
            js = e.get("judge_scores", {}).get(j)
            if js is not None:
                if isinstance(js, list):
                    scores.extend(js)
                else:
                    scores.append(js)
        if scores:
            all_scores.extend(scores)
            active_judges.append(j)

    if not all_scores:
        return {}

    arr = np.array(all_scores)
    mean = float(arr.mean())
    std = float(arr.std(ddof=1)) if len(arr) > 1 else 0.0

    # Histogram for distribution chart
    histogram = []
    if std > 0 and len(arr) >= 4:
        n_bins = min(20, max(8, int(np.sqrt(len(arr)))))
        counts, bin_edges = np.histogram(arr, bins=n_bins)
        histogram = [
            {"bin_center": round(float((bin_edges[i] + bin_edges[i + 1]) / 2), 2), "count": int(counts[i])}
            for i in range(len(counts))
        ]

    return {
        "mean": round(mean, 3),
        "std": round(std, 3),
        "total_judges": len(active_judges),
        "total_scores": len(all_scores),
        "histogram": histogram,
    }


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------

def build_summary(entries: list, judges: list, division_stats: dict) -> dict:
    # Flatten all division stats to find highlights
    all_rows = pd.concat(division_stats.values(), ignore_index=True) if division_stats else pd.DataFrame()

    def pick(col, ascending):
        if all_rows.empty or col not in all_rows.columns:
            return None
        row = all_rows.sort_values(col, ascending=ascending).iloc[0]
        return str(row["Judge_Full_Name"])

    divisions = list({e.get("division", "") for e in entries if e.get("division")})

    return {
        "total_competitors": len(entries),
        "total_judges": len(judges),
        "divisions": divisions,
        "judge_names": judges,
        "most_generous_judge": pick("Z_Severity_Index", ascending=False),
        "most_strict_judge": pick("Z_Severity_Index", ascending=True),
        "most_consistent_judge": pick("Absolute_Severity", ascending=True),
        "least_consistent_judge": pick("Absolute_Severity", ascending=False),
    }


# ---------------------------------------------------------------------------
# Excel Export — per-division sheets
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill(start_color="1B3A6B", end_color="1B3A6B", fill_type="solid")
HIGH_FILL    = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
WARN_FILL    = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
NORMAL_FILL  = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
DATA_FONT    = Font(color="000000", name="Calibri", size=11)
CENTER       = Alignment(horizontal="center", vertical="center", wrap_text=False)


def _format_p_value(val) -> str:
    """Format p-value as uppercase scientific notation: '3.41E-04'"""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return "N/A"
    return f"{val:.2E}"


def _auto_width(ws):
    """Set each column to fit its widest cell."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                max_len = max(max_len, cell_len)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max_len + 4


def export_excel(division_stats: dict, output_path: str):
    """
    Write one sheet per division. Each sheet has the 10 required columns
    with navy header, severity-based row color-coding, and frozen header row.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for div_name, df in division_stats.items():
        ws = wb.create_sheet(title=div_name[:31])  # Excel sheet name max 31 chars
        ws.freeze_panes = "A2"

        if df.empty:
            ws.cell(row=1, column=1, value=f"No data for division: {div_name}")
            continue

        # Header row
        for col_idx, col_name in enumerate(STAT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER

        # Data rows
        for row_offset, (_, row) in enumerate(df.iterrows()):
            r = row_offset + 2
            abs_sev = row.get("Absolute_Severity") or 0.0
            if abs_sev > 1.5:
                row_fill = HIGH_FILL
            elif abs_sev > 1.0:
                row_fill = WARN_FILL
            else:
                row_fill = NORMAL_FILL

            for col_idx, col_name in enumerate(STAT_COLUMNS, start=1):
                raw = row.get(col_name)

                # Format p_value as scientific notation string
                if col_name == "p_value":
                    display = _format_p_value(raw)
                elif col_name == "Mann_Whitney_U" and raw is None:
                    display = "N/A"
                elif raw is None or (isinstance(raw, float) and math.isnan(raw)):
                    display = "N/A"
                else:
                    display = raw

                cell = ws.cell(row=r, column=col_idx, value=display)
                cell.fill = row_fill
                cell.font = DATA_FONT
                cell.alignment = CENTER

        _auto_width(ws)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# JSON helpers
# ---------------------------------------------------------------------------

def _safe(v):
    if isinstance(v, float) and math.isnan(v):
        return None
    if isinstance(v, (np.integer,)):
        return int(v)
    if isinstance(v, (np.floating,)):
        return float(v)
    return v


def _df_to_list(df: pd.DataFrame) -> list:
    return [{k: _safe(v) for k, v in row.items()} for _, row in df.iterrows()]


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def analyze(input_path: str, output_dir: str) -> dict:
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    data, format_info = load_data(input_path)
    judge_id_map = format_info.get("judge_id_map", {})

    judges = extract_judge_names(data)

    # Per-division stats
    divisions = list({e.get("division", "") for e in data if e.get("division")})
    division_stats = {}
    division_stats_list = {}  # for JSON output
    division_summary = {}
    for div in divisions:
        div_entries = [e for e in data if e.get("division") == div]
        div_judges = [j for j in judges if any(
            e.get("judge_scores", {}).get(j) for e in div_entries
        )]
        div_df = compute_division_judge_stats(div_entries, div_judges, judge_id_map)
        division_stats[div] = div_df
        division_stats_list[div] = _df_to_list(div_df)
        division_summary[div] = compute_division_summary(div_entries, div_judges)

    summary = build_summary(data, judges, division_stats)

    excel_path = os.path.join(output_dir, "judgeiq_analytics.xlsx")
    export_excel(division_stats, excel_path)

    # Strip judge_id_map from format_info before returning (too verbose for response)
    format_info_clean = {k: v for k, v in format_info.items() if k != "judge_id_map"}

    return {
        "format_info": format_info_clean,
        "summary": summary,
        "division_stats": division_stats_list,
        "division_summary": division_summary,
        "excel_path": excel_path,
    }


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python analyze.py input.json output_dir/", file=sys.stderr)
        sys.exit(1)

    result = analyze(sys.argv[1], sys.argv[2])
    print(json.dumps(result, indent=2, default=str))
